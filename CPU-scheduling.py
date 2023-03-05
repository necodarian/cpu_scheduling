import abc
import time
from dataclasses import dataclass, field
from enum import Enum, auto
from typing import List

from openpyxl import load_workbook


@dataclass
class Job:
    arrival_time: int
    process_number: int
    burst_time: int
    priority: int

    remaining_time: int = field(init=False)
    waited_time: int = 0

    def __post_init__(self):
        self.remaining_time = self.burst_time

    def wait(self):
        self.waited_time += 1

    def run(self) -> None:
        self.remaining_time -= 1
        time.sleep(0.001)


def read_job_infos() -> List[Job]:
    wb = load_workbook(filename="cpu-scheduling.xlsx")
    ws = wb.active
    job_data = ws.iter_rows(min_row=2, max_row=21, min_col=1, max_col=4, values_only=True)

    jobs = []
    for job_info in job_data:
        jobs.append(
            Job(arrival_time=job_info[1], burst_time=job_info[2], process_number=job_info[0], priority=job_info[3]))
    return jobs


class Algorithms(Enum):
    FirstInFirstOut = auto()
    ShortestJobFirst = auto()
    Priority = auto()
    RoundRobin = auto()


class Algorithm(abc.ABC):
    def __init__(self, jobs: List[Job]):
        self.jobs = jobs

    @staticmethod
    def status(queue: List[Job], elapsed_time: int, current_job: Job) -> None:
        if current_job.remaining_time == 0:
            smg = f"Last instruction"
        else:
            smg = f"{current_job.remaining_time} instructions left"
        print(f"# Time Unit {elapsed_time}: PID {current_job.process_number} executes. "
              f"{smg}. Q={len(queue)}.")

        msg = ""
        for job in queue:
            msg += f"PID {job.process_number} wait={job.waited_time}. "
        if msg:
            print(msg)

    @abc.abstractmethod
    def run(self):
        pass


class FirstInFirstOut(Algorithm):
    def __init__(self, jobs: List[Job]):
        super().__init__(jobs)

    def run(self):
        if not self.jobs:
            return

        first_job = next(iter(self.jobs))
        elapsed_time = 0
        remaining_jobs = [job for job in self.jobs]

        queue = [job for job in self.jobs if job.arrival_time == first_job.arrival_time]
        while queue:
            running_job = queue.pop(0)
            remaining_jobs.remove(running_job)

            for i in range(running_job.burst_time):
                elapsed_time += 1
                running_job.run()
                for job in queue:
                    job.wait()

                new_jobs = [j for j in remaining_jobs if j.arrival_time <= elapsed_time and j not in queue]
                for j in new_jobs:
                    queue.append(j)
                    j.wait()

                self.status(queue, elapsed_time, running_job)

            if queue:
                elapsed_time += 1
                print(f"# Time Unit {elapsed_time}: Context switch.")
                msg = ""
                for job in queue[1:]:
                    job.wait()
                    msg += f"PID {job.process_number} wait={job.waited_time}. "
                if msg:
                    print(msg)


class ShortestJobFirst(Algorithm):
    def __init__(self, jobs: List[Job]):
        super().__init__(jobs)

    @staticmethod
    def sort(queue: List[Job]) -> List[Job]:
        queue.sort(key=lambda j: j.burst_time)

        tmp = {j.burst_time: {} for j in queue}

        for j in queue:
            if tmp[j.burst_time].get(j.arrival_time, None) is None:
                tmp[j.burst_time][j.arrival_time] = []
            tmp[j.burst_time][j.arrival_time].append(j)
            tmp[j.burst_time][j.arrival_time].sort(key=lambda y: y.process_number)
            tmp[j.burst_time] = dict(sorted(tmp[j.burst_time].items()))

        queue = []
        for a in tmp.values():
            for b in a.values():
                for j in b:
                    queue.append(j)
        return queue

    def run(self):
        if not self.jobs:
            return

        first_job = next(iter(self.jobs))
        elapsed_time = 0
        remaining_jobs = [job for job in self.jobs]

        queue = sorted([job for job in self.jobs if job.arrival_time == first_job.arrival_time],
                       key=lambda j: j.burst_time)
        while queue:
            running_job = queue.pop(0)
            remaining_jobs.remove(running_job)

            for i in range(running_job.burst_time):
                elapsed_time += 1
                running_job.run()
                for job in queue:
                    job.wait()

                new_jobs = [j for j in remaining_jobs if j.arrival_time <= elapsed_time and j not in queue]
                for j in new_jobs:
                    queue.append(j)
                    j.wait()

                queue = self.sort(queue)
                self.status(queue, elapsed_time, running_job)

            if queue:
                elapsed_time += 1
                print(f"# Time Unit {elapsed_time}: Context switch.")
                msg = ""
                for job in queue[1:]:
                    job.wait()
                    msg += f"PID {job.process_number} wait={job.waited_time}. "
                if msg:
                    print(msg)


class Priority(Algorithm):
    def __init__(self, jobs: List[Job]):
        super().__init__(jobs)

    @staticmethod
    def sort(queue: List[Job]) -> List[Job]:
        queue.sort(key=lambda j: j.priority)

        tmp = {j.priority: {} for j in queue}
        for j in queue:
            if tmp[j.priority].get(j.arrival_time, None) is None:
                tmp[j.priority][j.arrival_time] = []
            tmp[j.priority][j.arrival_time].append(j)
            tmp[j.priority][j.arrival_time].sort(key=lambda y: y.process_number)
            tmp[j.priority] = dict(sorted(tmp[j.priority].items()))

        tmp = dict(sorted(tmp.items()))

        queue = []
        for a in tmp.values():
            for b in a.values():
                for j in b:
                    queue.append(j)

        return queue

    def run(self):
        if not self.jobs:
            return

        first_job = next(iter(self.jobs))
        elapsed_time = 0
        remaining_jobs = [job for job in self.jobs]

        queue = [job for job in self.jobs if job.arrival_time == first_job.arrival_time]
        queue = self.sort(queue)
        while queue:
            running_job = queue.pop(0)

            while running_job.remaining_time > 0:
                elapsed_time += 1
                running_job.run()

                new_jobs = [j for j in remaining_jobs
                            if j.arrival_time <= elapsed_time and j not in queue and j is not running_job]
                for j in new_jobs:
                    queue.append(j)

                for job in queue:
                    job.wait()
                queue = self.sort(queue)
                self.status(queue, elapsed_time, running_job)

            remaining_jobs.remove(running_job)

            if queue:
                elapsed_time += 1
                print(f"# Time Unit {elapsed_time}: Context switch.")
                msg = ""
                for job in queue[1:]:
                    job.wait()
                    msg += f"PID {job.process_number} wait={job.waited_time}. "
                if msg:
                    print(msg)


class RoundRobin(Algorithm):
    def __init__(self, jobs: List[Job], slice_time: int = 4):
        super().__init__(jobs)
        self.slice_time = slice_time

    @staticmethod
    def sort(queue: List[Job]) -> List[Job]:
        queue.sort(key=lambda j: j.process_number)

        tmp = {j.arrival_time: [] for j in queue}
        for j in queue:
            tmp[j.arrival_time].append(j)
            tmp[j.arrival_time].sort(key=lambda y: y.burst_time)

        queue = []
        for a in tmp.values():
            for j in a:
                queue.append(j)

        return queue

    def run(self):
        if not self.jobs:
            return

        elapsed_time = 0
        remaining_jobs = [job for job in self.jobs]

        queue = [job for job in self.jobs]
        queue = self.sort(queue)

        while remaining_jobs:
            running_job = queue.pop(0)

            number_of_time_units = 0
            while running_job.remaining_time > 0:
                elapsed_time += 1
                number_of_time_units += 1
                running_job.run()

                new_jobs = [j for j in remaining_jobs if j.arrival_time <= elapsed_time and j is not running_job]
                for job in new_jobs:
                    job.wait()

                self.status(new_jobs, elapsed_time, running_job)

                if number_of_time_units >= self.slice_time:
                    if running_job.remaining_time > 0:
                        queue.append(running_job)
                    else:
                        remaining_jobs.remove(running_job)
                    running_job = queue.pop(0)
                    number_of_time_units = 0

            remaining_jobs.remove(running_job)

            if queue:
                elapsed_time += 1
                print(f"# Time Unit {elapsed_time}: Context switch.")
                msg = ""
                for job in queue[1:]:
                    job.wait()
                    msg += f"PID {job.process_number} wait={job.waited_time}. "
                if msg:
                    print(msg)


def main():
    print("Welcome to CPU Schedular")
    print("The following algorithms are available")
    for algo in Algorithms:
        print(f"{algo.value}. {algo.name}")

    selected_algo_index = int(input("Select a number: "))
    if selected_algo_index == 1:
        algo = FirstInFirstOut
    elif selected_algo_index == 2:
        algo = ShortestJobFirst
    elif selected_algo_index == 3:
        algo = Priority
    elif selected_algo_index == 4:
        algo = RoundRobin
    else:
        raise RuntimeError("Invalid selection. Try again.")

    jobs = read_job_infos()
    a = algo(jobs)
    a.run()


if __name__ == '__main__':
    main()
